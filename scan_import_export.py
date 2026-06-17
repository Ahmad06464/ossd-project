import json
import re
from datetime import datetime, timezone
from pathlib import Path

EXPORT_VERSION = 1
MARKER = "RECON SCAN EXPORT v1"
JSON_BEGIN = "RECON_JSON_BEGIN"
JSON_END = "RECON_JSON_END"


def build_scan_payload(
    target,
    results,
    profile="",
    scan_completed_at="",
    targets=None,
    exported_at=None,
    resume_fields=None,
):
    payload = {
        "version": EXPORT_VERSION,
        "target": target or "",
        "targets": targets or ([target] if target else []),
        "profile": profile or "",
        "scan_completed_at": scan_completed_at or "",
        "exported_at": exported_at or datetime.now(timezone.utc).isoformat(),
        "results": results,
    }
    if resume_fields:
        payload.update(resume_fields)
    return payload


def format_scan_time(iso_str):
    if not iso_str:
        return "Unknown time"
    try:
        dt = datetime.fromisoformat(str(iso_str).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    except (TypeError, ValueError):
        return str(iso_str)


def embed_json_block(payload):
    return f"{JSON_BEGIN}\n{json.dumps(payload, indent=2)}\n{JSON_END}"


def extract_json_block(text):
    start = text.find(JSON_BEGIN)
    end = text.find(JSON_END)
    if start == -1 or end == -1 or end <= start:
        return None
    block = text[start + len(JSON_BEGIN):end].strip()
    return json.loads(block)


def normalize_import_data(data):
    if not isinstance(data, dict):
        raise ValueError("Invalid scan export format.")
    if "results" not in data:
        raise ValueError("Scan export has no results.")
    return data


def import_scan_file(path):
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".json":
        with open(path, encoding="utf-8") as handle:
            return normalize_import_data(json.load(handle))

    if suffix == ".txt":
        return normalize_import_data(_parse_txt_export(path.read_text(encoding="utf-8", errors="ignore")))

    if suffix in (".xlsx", ".xlsm"):
        return normalize_import_data(_parse_excel_export(path))

    if suffix == ".html":
        return normalize_import_data(_parse_html_export(path.read_text(encoding="utf-8", errors="ignore")))

    raise ValueError(f"Unsupported file type: {suffix}")


def _parse_txt_export(text):
    data = extract_json_block(text)
    if data:
        return data
    raise ValueError("This TXT file is not a Recon scan export (missing RECON_JSON block).")


def _parse_html_export(text):
    match = re.search(
        r'<script\s+type="application/json"\s+id="recon-scan-export"\s*>\s*(.*?)\s*</script>',
        text,
        re.DOTALL | re.IGNORECASE,
    )
    if match:
        return json.loads(match.group(1))
    data = extract_json_block(text)
    if data:
        return data
    raise ValueError("This HTML file is not a Recon scan export.")


def _parse_excel_export(path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "ScanMeta" in workbook.sheetnames:
            raw = workbook["ScanMeta"]["A1"].value
            if raw:
                return json.loads(str(raw))
    finally:
        workbook.close()

    raise ValueError("This Excel file is not a Recon scan export (missing ScanMeta sheet).")
